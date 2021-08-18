# -*- coding: utf-8 -*-
"""
Created on Thu Aug  5 15:27:18 2021

@author: AmaliTech

For creating master files

====================================
This module is for automatically creating mater files 
corresponding to the courses available

1. read the courselist.csv file
2. for each course in the course list, creat a master excel file
    and save to master folder
"""
from openpyxl import load_workbook,Workbook
import pandas as pd
import os
from common import bulk_load_data
from common import TRACK_COURSE_MAPPING,MASTER_FOLDER

#master_folder = os.path.join(os.getcwd(),'master')

course_df = pd.read_csv('courselist.csv')
all_courses = [ name.strip() for name in course_df.name.values]


namesfile = 'names.csv'

#get the people enrolled
names_df = pd.read_csv(namesfile)
names_df['name'] = names_df['name'].apply(lambda x: x.strip()) 

trainee_list = names_df.to_dict(orient='record')
num_enrolled = len(trainee_list)

# create a workbook 

def create_single_master():
    """
    Create one csv file containing progress reports for all trainees for all
    courses

    Returns
    -------
    .csv file
    """
    wb = Workbook()
    summary = wb.active
    summary.title = 'summary'
    progress = wb.create_sheet(title = 'progress')
    
    # Create Sheet Headers
    # Create the sheet headers
    summary.cell(row = 1,column = 1).value = 'Name'
    summary.cell(row = 1,column = 2).value = 'Track'
    summary.cell(row = 1,column = 3).value = 'Course'
    summary.cell(row = 1,column = 4).value = 'StartedAt'
    summary.cell(row = 1,column = 5).value = 'CompletedAt'
    summary.cell(row = 1,column = 6).value = 'TimeTaken(Days)'
    
    
    # Create progression sheet headers
    progress.cell(row = 1,column = 1).value = 'Name'
    progress.cell(row = 1,column = 2).value = 'Track'
    progress.cell(row = 1,column = 3).value = 'Course'
        
    # populate summary sheet with data
    for trainee in trainee_list:
        # determine the last row for data insertion
        #last_row = summary.max_row
        course_list = TRACK_COURSE_MAPPING.get(trainee['track'])
        
        if course_list:
            for course in course_list.values():
                #Populate summary sheet
                row = summary.max_row + 1
                summary.cell(row = row, column = 1).value = trainee['name']
                summary.cell(row = row, column = 2).value = trainee['track']
                summary.cell(row = row, column = 3).value = course
                summary.cell(row = row, column = 6).value = "=DAYS(E{},D{})".format(row,row)
                
                # Populate progress Sheet
                progress.cell(row = row, column = 1).value = trainee['name']
                progress.cell(row = row, column = 2).value = trainee['track']
                progress.cell(row = row, column = 3).value = course
        else:
            continue
        
        #jumpt_to_row = idx + 1+ len(course_list)
    
    
    # save file
    output_file = os.path.join(MASTER_FOLDER,'course_progress' +'.xlsx')
    if not os.path.exists(output_file):
        wb.save(output_file)
    else:

        response = input('The output file already exists at the destiantion. Overwrite? (Y/N):  ')
        if response.lower() == 'y':
            wb.save(output_file)
        

def create_seperate_masters(course_list = all_courses,name_list = trainee_list):
    for course in all_courses:
        wb = Workbook()
        summary = wb.active
        summary.title = 'summary'
        progress = wb.create_sheet(title = 'progress')
        
        # Create the sheet headers in row 1
        summary.cell(row = 1,column = 1).value = 'Name'
        summary.cell(row = 1,column = 2).value = 'Track'
        summary.cell(row = 1,column = 3).value = 'Course'
        summary.cell(row = 1,column = 4).value = 'StartedAt'
        summary.cell(row = 1,column = 5).value = 'CompletedAt'
        summary.cell(row = 1,column = 6).value = 'TimeTaken(Days)'
        
        # Create progression sheet headers
        progress.cell(row = 1,column = 1).value = 'Name'
        progress.cell(row = 1,column = 2).value = 'Track'
        progress.cell(row = 1,column = 3).value = 'Course'
            
        
        
        # Get people enrolled in course & create entries for summary sheet
                
        for row, person in enumerate(name_list,2):
                
            name, email, track = person
            try:
                track_courses = TRACK_COURSE_MAPPING.get(person[track],None).values()
                
                summary.cell(row = row, column = 1).value = person[name]
                summary.cell(row = row, column = 2).value = person[track]
                summary.cell(row = row, column = 3).value = course
                
                summary.cell(row = row, column = 6).value = "=DAYS(E{},D{})".format(row,row)
            
                # create entries for progress sheet
                progress.cell(row = row, column = 1).value = person[name]
                progress.cell(row = row, column = 2).value = person[track]
                progress.cell(row = row, column = 3).value = course
                
            except AttributeError:
                #print('{} Course not found for {}'.format(course,person[name]))
                continue
                
           
            #if this person is enrolled in this course,add their detail
                
                 
        
        # save file
        output_file = os.path.join(MASTER_FOLDER,course +'.xlsx')
        #if not os.path.exists(output_file):
        wb.save(output_file) 


    
    
    
    
if __name__ == 'main':
    create_seperate_masters()
else:
    create_seperate_masters()
    
    

