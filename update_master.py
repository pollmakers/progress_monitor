# -*- coding: utf-8 -*-
"""
Created on Thu Jul 29 13:57:14 2021

@author: AmaliTech

=====================================
This module is for updatiing a single master sheet with the weekly updates.

1. read the progress file
2. select the percent_complete column
3. append the data to the coresponding name in the master sheet.

NB: Need to make sure that names are correctly matched
"""
import datetime
from openpyxl import load_workbook
import pandas as pd
import os
import glob 
import shutil

from common import INPUT_FOLDER, OUTPUT_FOLDER, COURSE_MAPPING_FILE, NAMES_FILE, TRACK_COURSE_MAPPING
from common import MASTER_FOLDER, HISTORY_FOLDER

from common import bulk_load_data,get_progress_data, load_data
#========= Test Data ===
progress_file = r"C:\Users\AmaliTech\Documents\CourseProgress Processing\output\current_progress.csv"
master_workbook = load_workbook(os.path.join(MASTER_FOLDER,'course_progress.xlsx'))
#=======================

# list available courses
course_df = pd.read_csv('courselist.csv')
namesfile = 'names.csv'

#get the people enrolled
names_df = pd.read_csv(namesfile)
names_df['name'] = names_df['name'].apply(lambda x: x.strip()) 

trainee_list = names_df.to_dict(orient='record')
num_enrolled = len(trainee_list)

psheet = master_workbook['progress']

def get_master_data():
    try:
        #master_file = os.path.basename(file_name.split('.')[0]) #+ ".xlsx"
        master_file = os.path.join(MASTER_FOLDER,'course_progress.xlsx')
        master_wb = load_workbook(master_file)
    except FileNotFoundError:
        print("{} does not exist in master folder".format(master_file))
        return
    
    return master_file,master_wb

def update_progress():
    """
    obtains the progress info for each trainee for each course enroled and 
    updates the corresponding records

    Returns
    -------
    str
        updated version of progress workbok in master folder.

    """
    # obtain the sheet to be updated
    master_progress_sheet = master_workbook['progress']
    master_summary_sheet = master_workbook['summary']
    
    # make sheets in df
    progress_columns = next(master_progress_sheet.values)
    master_progress_sheet_df = pd.DataFrame(
        list(master_progress_sheet.values),
        columns=progress_columns
        )
    last_column = master_progress_sheet.max_column
    next_column = last_column + 1
    
    #create a new column with header as current date
    master_progress_sheet.cell(1,next_column).value = datetime.datetime.today().date()
    
    # create dictionary mapings to ease out name search
    
    
    # go through the progress files and update the scores for all
    # enrolled in that course
    for track in TRACK_COURSE_MAPPING:
        course_list = TRACK_COURSE_MAPPING.get(track)
        progress_files = [os.path.join(os.getcwd(),OUTPUT_FOLDER,f) for f in course_list.keys()]
        
        for file in progress_files:
            
            
            progress_data = load_data(file)
            
            # get the people enroled in this course:
            course = os.path.basename(file)
            people_enrolled_data = progress_data[(progress_data.course == course) & (progress_data.enrolled == 'Yes')]
            
            name_progress_dict = { name.strip():progress for name,progress in list(people_enrolled_data[['name','percent_complete']].values)}
            name_date_started_dict = { name.strip():started for name,started in list(people_enrolled_data[['name','started_at']].values)}
            name_date_completed_dict = { name.strip():started for name,started in list(people_enrolled_data[['name','completed_at']].values)}
    
            #Now conver master progress sheet to pandas df, apply neccesaey filter and update
            #percent_complete = progress_data.percent_complete
            for row in range(2, master_progress_sheet.max_row+1):
                
                # get the name of current learner
                
                name = master_progress_sheet.cell(row = row, column=1).value.strip()
                
                # find corresponding value and update the progress
                
                master_progress_sheet.cell(row = row, column=next_column).value = name_progress_dict.get(name,0)
                
                # if Started_At column is empty, update it
                if master_summary_sheet.cell(row = row, column=3).value is None:
                   master_summary_sheet.cell(row = row, column=3).value = name_date_started_dict.get(name,'')
                   master_summary_sheet.cell(row = row, column=4).value = name_date_completed_dict.get(name,'')
            #return master_progress_sheet
        
    

    

def update_progress_sheet(progress_data,master_workbook):
    """
    This function handles the task of updating the master sheet with the 
    latest progress figures.
    
    It creates a new column with with the date the progress data is read
    and add the progress reading
    
    To Do:
        Add a function that checks and update the Started At column
    """
    _,master_workbook = get_master_data()
    master_progress_sheet = master_workbook['progress']
    summary_sheet = master_workbook['summary']
    last_column = master_progress_sheet.max_column
    next_column = last_column + 1
    
    #create a new column with header as current date
    master_progress_sheet.cell(1,next_column).value = datetime.datetime.today().date()
    
    # create dictionary mapings to ease out name search
    
    name_progress_dict = { name.strip():progress for name,progress in list(progress_data[['name','percent_complete']].values)}
    name_date_started_dict = { name.strip():started for name,started in list(progress_data[['name','started_at']].values)}
    name_date_completed_dict = { name.strip():started for name,started in list(progress_data[['name','completed_at']].values)}
    
    #now fill the remaining cells with the corresponding values
    #percent_complete = progress_data.percent_complete
    for row in range(2, master_progress_sheet.max_row+1):
        
        # get the name of current learner
        
        name = master_progress_sheet.cell(row = row, column=1).value.strip()
        
        # find corresponding value and update the progress
        
        master_progress_sheet.cell(row = row, column=next_column).value = name_progress_dict.get(name,0)
        
        # if Started_At column is empty, update it
        if summary_sheet.cell(row = row, column=3).value is None:
           summary_sheet.cell(row = row, column=3).value = name_date_started_dict.get(name,'')
           summary_sheet.cell(row = row, column=4).value = name_date_completed_dict.get(name,'')
    #return master_progress_sheet


def cleanup():
    """
    Deletes all the progress files from the input and output folders
    - input files are backed up in the history folder with subfolders by the date of update
    - files in output folders are deleted completely
    
    1. Get updated progress files
    2. Move them to history
    """
    current_date = str(datetime.datetime.today().date())
    
    # load progress files used
    progress_files_folder = os.path.join(os.getcwd(),'input')
    updated_files = bulk_load_data(progress_files_folder)
    
    # create a new folder and move items into it
    destination_folder = os.path.join(HISTORY_FOLDER,current_date)
    
    if not os.path.exists(destination_folder):
        os.mkdir(destination_folder)
        
    for file in updated_files:
        file_name = os.path.basename(file)
        shutil.move(file,os.path.join(destination_folder,file_name))
        
    # Cleanup files in the output folder in parent directory
    for file in bulk_load_data(INPUT_FOLDER):
        os.remove(file)
    
    
    
    
    
    
    
"""
TO DO
- Add function to add the date started in the summary sheet each time 
master file is updated for those leaeners whose entries are empty
"""    

#wb = load_workbook(filename = r"C:\Users\AmaliTech\Documents\CourseProgress Processing\master\build-web-apps-with-django.xlsx")
#sheet = wb.get_sheet_by_name('progression')

    


# Obtain the new progress data. Column of interest is current_progress
#progress_data = load_data(progress_file)




# def process(mode = 'single'):
#     response = input("Do you want to process multiple files? Y/N:  ")
    
#     mode = modes[response.lower()]
#     if mode == 'single':
#         print()
#         print("Make sure the file specified is in the `input` folder")
#         print("The output will be in the `output` folder with the same name")
#         file_name = input("What is the name of the progress file without the extension?   ")
        
#         # 1. Get progress Data
#         progress_file, progress_data = get_progress_data(file_name = file_name )
        
#         # 2. Get master file
#         master_file, master_wb = get_master_data(file_name = progress_file) 
        
#         # 3. update progress sheet in master workbook
#         update_progress_sheet(progress_data = progress_data,
#                               master_workbook= master_wb
#                               )
        
#         # 4. Save workbook
#         master_wb.save(os.path.join(MASTER_FOLDER,master_file + '_copy.xlsx'))
#         print("Done updating **{}** in master  folder".format(master_file))
        

    
    
#         #master_wb.save(os.path.join(MASTER_FOLDER,master_file + '_copy.xlsx'))
#     elif mode == 'bulk':
#         bulk_files = bulk_load_data()
        
#         for file in bulk_files:
#             # 1. Get progress Data
#             progress_file, progress_data = get_progress_data(file_name = file )
        
#             try:
#                 # 2. Get master file
#                 master_file, master_wb = get_master_data(file_name = progress_file)
               
            
                
                
#                # 3. update progress sheet in master workbook
#                 update_progress_sheet(progress_data = progress_data,
#                                   master_workbook= master_wb,
#                                   )
                
#                 # 4. Save workbook
#                 master_wb.save(os.path.join(MASTER_FOLDER,master_file + '_copy.xlsx'))
#                 print("Done updating **{}** in master  folder".format(master_file))
        
#             except TypeError:
#                 continue
        
#         cleanup()


if __name__ == 'main':
    update_progress()
else:
    update_progress()
    